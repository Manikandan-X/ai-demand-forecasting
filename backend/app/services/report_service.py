import os
import pandas as pd

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer
)

from reportlab.lib.styles import (
    getSampleStyleSheet
)

from openpyxl import Workbook
from openpyxl.styles import Font

from fastapi import HTTPException

from sqlalchemy.orm import Session

from app.models.dataset import Dataset

from app.services.notification_service import (
    NotificationService
)

from app.utils.activity_logger import (
    log_user_activity
)

from app.routers.websocket import (
    send_dashboard_update
)


REPORT_FOLDER = "reports"

os.makedirs(
    REPORT_FOLDER,
    exist_ok=True
)


# ANALYTICS SUMMARY
def generate_analytics_summary(df):

    summary = {

        "total_sales":
            float(
                df["Total_Amount"].sum()
            ),

        "total_orders":
            int(len(df)),

        "top_product":
            df["Product"]
            .value_counts()
            .idxmax(),

        "average_order_value":
            float(
                df["Total_Amount"]
                .mean()
            )
    }

    return summary


# EXCEL REPORT
def generate_excel_report(
    df,
    filename
):

    report_path = os.path.join(
        REPORT_FOLDER,
        filename
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Forecast Report"

    # TITLE
    sheet["A1"] = (
        "AI Forecast Analytics Report"
    )

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    summary = generate_analytics_summary(df)

    # SUMMARY SECTION
    sheet["A3"] = "Total Sales"
    sheet["B3"] = summary[
        "total_sales"
    ]

    sheet["A4"] = "Total Orders"
    sheet["B4"] = summary[
        "total_orders"
    ]

    sheet["A5"] = "Top Product"
    sheet["B5"] = summary[
        "top_product"
    ]

    sheet["A6"] = (
        "Average Order Value"
    )

    sheet["B6"] = summary[
        "average_order_value"
    ]

    # DATA HEADER
    start_row = 9

    for col_num, column_name in enumerate(
        df.columns,
        1
    ):

        cell = sheet.cell(
            row=start_row,
            column=col_num
        )

        cell.value = column_name

        cell.font = Font(
            bold=True
        )

    # DATA ROWS
    for row_index, row in enumerate(
        df.values,
        start_row + 1
    ):

        for col_index, value in enumerate(
            row,
            1
        ):

            sheet.cell(
                row=row_index,
                column=col_index,
                value=str(value)
            )

    workbook.save(report_path)

    return report_path


# PDF REPORT
def generate_pdf_report(
    df,
    filename
):

    report_path = os.path.join(
        REPORT_FOLDER,
        filename
    )

    document = SimpleDocTemplate(
        report_path
    )

    styles = getSampleStyleSheet()

    elements = []

    # TITLE
    title = Paragraph(
        "AI Forecast Analytics Report",
        styles["Title"]
    )

    elements.append(title)

    elements.append(
        Spacer(1, 20)
    )

    summary = generate_analytics_summary(df)

    # SUMMARY SECTION
    elements.append(

        Paragraph(
            f"Total Sales: "
            f"{summary['total_sales']}",

            styles["BodyText"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    elements.append(

        Paragraph(
            f"Total Orders: "
            f"{summary['total_orders']}",

            styles["BodyText"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    elements.append(

        Paragraph(
            f"Top Product: "
            f"{summary['top_product']}",

            styles["BodyText"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    elements.append(

        Paragraph(
            f"Average Order Value: "
            f"{summary['average_order_value']}",

            styles["BodyText"]
        )
    )

    elements.append(
        Spacer(1, 20)
    )

    # SAMPLE DATA
    sample_title = Paragraph(
        "Sample Dataset Records",
        styles["Heading2"]
    )

    elements.append(sample_title)

    elements.append(
        Spacer(1, 10)
    )

    for _, row in df.head(10).iterrows():

        row_text = Paragraph(
            str(row.to_dict()),
            styles["BodyText"]
        )

        elements.append(row_text)

        elements.append(
            Spacer(1, 8)
        )

    document.build(elements)

    return report_path

class ReportService:

    @staticmethod
    async def export_excel_report(

        db: Session,

        dataset_id: int,

        current_user
    ):

        dataset = (

            db.query(Dataset)

            .filter(
                Dataset.id == dataset_id,

                Dataset.uploaded_by ==
                current_user.id
            )

            .first()
        )

        if not dataset:

            raise HTTPException(
                status_code=404,
                detail="Dataset not found"
            )

        if dataset.file_path.endswith(".csv"):

            df = pd.read_csv(
                dataset.file_path
            )

        else:

            df = pd.read_excel(
                dataset.file_path
            )

        report_path = (
            generate_excel_report(
                df,
                f"dataset_{dataset_id}.xlsx"
            )
        )

        log_user_activity(

            db=db,

            user_id=current_user.id,

            action=
            "REPORT_EXCEL_DOWNLOAD",

            details=(
                f"Excel report "
                f"downloaded for "
                f"{dataset.filename}"
            )
        )

        NotificationService.create_notification(

            db=db,

            title=
            "Report Generated",

            message=
            "Excel report generated successfully",

            user_id=
            current_user.id,

            notification_type=
            "report",

            is_admin=False
        )

        await send_dashboard_update(
            user_id=current_user.id
        )

        return report_path

    @staticmethod
    async def export_pdf_report(

        db: Session,

        dataset_id: int,

        current_user
    ):

        dataset = (

            db.query(Dataset)

            .filter(
                Dataset.id == dataset_id,

                Dataset.uploaded_by ==
                current_user.id
            )

            .first()
        )

        if not dataset:

            raise HTTPException(
                status_code=404,
                detail="Dataset not found"
            )

        if dataset.file_path.endswith(".csv"):

            df = pd.read_csv(
                dataset.file_path
            )

        else:

            df = pd.read_excel(
                dataset.file_path
            )

        report_path = (
            generate_pdf_report(
                df,
                f"dataset_{dataset_id}.pdf"
            )
        )

        log_user_activity(

            db=db,

            user_id=current_user.id,

            action=
            "REPORT_PDF_DOWNLOAD",

            details=(
                f"PDF report "
                f"downloaded for "
                f"{dataset.filename}"
            )
        )

        NotificationService.create_notification(

            db=db,

            title=
            "Report Generated",

            message=
            "PDF report generated successfully",

            user_id=
            current_user.id,

            notification_type=
            "report",

            is_admin=False
        )

        await send_dashboard_update(
            user_id=current_user.id
        )

        return report_path